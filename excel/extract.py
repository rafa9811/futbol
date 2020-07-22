import openpyxl


def get_teams():
    ps = openpyxl.load_workbook('teams_data.xlsx')
    sheet = ps['Hoja 1']
    teamlist = []
    for i in range(5, 45):
        teamlist.append(sheet['A'+str(i)].value)
    return teamlist


def team_row(team):
    ps = openpyxl.load_workbook('teams_data.xlsx')
    sheet = ps['Hoja 1']
    for i in range(5, 45):
        if sheet['A'+str(i)].value == team:
            return i
    return -1


def get_player_first_kit(team):
    ps = openpyxl.load_workbook('teams_data.xlsx')
    sheet = ps['Hoja 1']
    row = str(team_row(team))

    shirt1 = sheet['G'+row].fill.bgColor.rgb
    shirt2 = sheet['H'+row].fill.bgColor.rgb
    shorts1 = sheet['I'+row].fill.bgColor.rgb
    shorts2 = sheet['J'+row].fill.bgColor.rgb
    shocks1 = sheet['K'+row].fill.bgColor.rgb
    shocks2 = sheet['L'+row].fill.bgColor.rgb

    return {'shirt1':shirt1, 'shirt2':shirt2, 'shorts1':shorts1, 'shorts2':shorts2, 'shocks1':shocks1, 'shocks2':shocks2}


def get_gk_first_kit(team):
    ps = openpyxl.load_workbook('teams_data.xlsx')
    sheet = ps['Hoja 1']
    row = str(team_row(team))
    shirt1 = sheet['M'+row].fill.bgColor.rgb
    shirt2 = sheet['N'+row].fill.bgColor.rgb
    shorts1 = sheet['O'+row].fill.bgColor.rgb
    shorts2 = sheet['P'+row].fill.bgColor.rgb
    shocks1 = sheet['Q'+row].fill.bgColor.rgb
    shocks2 = sheet['R'+row].fill.bgColor.rgb

    return {'shirt1':shirt1, 'shirt2':shirt2, 'shorts1':shorts1, 'shorts2':shorts2, 'shocks1':shocks1, 'shocks2':shocks2}


def get_player_scnd_kit(team):
    ps = openpyxl.load_workbook('teams_data.xlsx')
    sheet = ps['Hoja 1']
    row = str(team_row(team))

    shirt1 = sheet['S'+row].fill.bgColor.rgb
    shirt2 = sheet['T'+row].fill.bgColor.rgb
    shorts1 = sheet['U'+row].fill.bgColor.rgb
    shorts2 = sheet['V'+row].fill.bgColor.rgb
    shocks1 = sheet['W'+row].fill.bgColor.rgb
    shocks2 = sheet['X'+row].fill.bgColor.rgb

    return {'shirt1':shirt1, 'shirt2':shirt2, 'shorts1':shorts1, 'shorts2':shorts2, 'shocks1':shocks1, 'shocks2':shocks2}


def get_gk_scnd_kit(team):
    ps = openpyxl.load_workbook('teams_data.xlsx')
    sheet = ps['Hoja 1']
    row = str(team_row(team))

    shirt1 = sheet['Y'+row].fill.bgColor.rgb
    shirt2 = sheet['Z'+row].fill.bgColor.rgb
    shorts1 = sheet['AA'+row].fill.bgColor.rgb
    shorts2 = sheet['AB'+row].fill.bgColor.rgb
    shocks1 = sheet['AC'+row].fill.bgColor.rgb
    shocks2 = sheet['AD'+row].fill.bgColor.rgb

    return {'shirt1':shirt1, 'shirt2':shirt2, 'shorts1':shorts1, 'shorts2':shorts2, 'shocks1':shocks1, 'shocks2':shocks2}
